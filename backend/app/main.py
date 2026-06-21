import os
import io
import tempfile
import uuid
import shutil
import glob
from pathlib import Path
from datetime import datetime, timezone, timedelta

BJT = timezone(timedelta(hours=8))
def now_bj() -> datetime:
    return datetime.now(BJT)

from fastapi import FastAPI, UploadFile, File, Form
from fastapi.responses import JSONResponse, FileResponse, StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles

from .storage import load_all, save_all
from .parser import parse_upload, confirm_upload as json_confirm_upload
from .database import db_session, Product, PriceHistory, upsert_price, get_product_id, get_product_name
from .migration import migrate_if_needed
from sqlalchemy import text as sa_text, func

app = FastAPI(title="Price Tracker")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

_preview_sessions: dict[str, dict] = {}

DATA_DIR = os.environ.get("DATA_DIR", "/data")
BACKUP_DIR = os.path.join(DATA_DIR, "backups")


# ── Startup hook ─────────────────────────────────────────────

@app.on_event("startup")
async def startup():
    """Run data migration at first start."""
    migrated = migrate_if_needed()
    if migrated:
        print("[startup] Migration complete, system ready")
    else:
        print("[startup] System ready")


# ── SQLite helpers ───────────────────────────────────────────

def _db_to_brand_groups():
    """Read from SQLite and return the same brand→items structure for backward compat."""
    from collections import defaultdict
    brands_map = defaultdict(list)
    with db_session() as session:
        products = session.query(Product).order_by(Product.id).all()
        for prod in products:
            records = session.query(PriceHistory).filter(
                PriceHistory.product_id == prod.id
            ).order_by(PriceHistory.price_date).all()
            prices = [{"date": r.price_date, "price": r.price} for r in records]
            brands_map[prod.brand or "未知"].append({
                "name": prod.name,
                "prices": prices,
            })
    brands = []
    for brand_name in sorted(brands_map.keys()):
        brands.append({"brand": brand_name, "items": brands_map[brand_name]})
    return brands


# ── Upload API (two-phase, writes to both SQLite + JSON) ─────

@app.post("/api/upload")
async def upload_legacy(file: UploadFile = File(...)):
    if not file.filename.endswith((".xlsx", ".xls")):
        return JSONResponse(status_code=400, content={"error": "Only .xlsx / .xls files accepted"})
    suffix = Path(file.filename).suffix
    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
        tmp.write(await file.read())
        tmp_path = tmp.name
    try:
        preview = parse_upload(tmp_path)
        result = json_confirm_upload(tmp_path, preview["date_str"], {})
        # Also write to SQLite
        _sync_json_to_sqlite(result)
        return {"status": "ok", "brands_count": len(result.get("brands", []))}
    finally:
        os.unlink(tmp_path)


@app.post("/api/upload/preview")
async def upload_preview(file: UploadFile = File(...)):
    if not file.filename.endswith((".xlsx", ".xls")):
        return JSONResponse(status_code=400, content={"error": "Only .xlsx / .xls files accepted"})
    suffix = Path(file.filename).suffix
    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
        tmp.write(await file.read())
        tmp_path = tmp.name
    try:
        result = parse_upload(tmp_path)
        session_id = uuid.uuid4().hex[:12]
        _preview_sessions[session_id] = {"filepath": tmp_path, "parsed_data": result["parsed_data"], "date_str": result["date_str"]}
        return {
            "session_id": session_id,
            "date_str": result["date_str"],
            "brands": result["parsed_data"].get("brands", []),
            "alerts": result["alerts"],
            "total_items": result["total_items"],
            "alert_count": len(result["alerts"]),
        }
    except Exception as e:
        os.unlink(tmp_path)
        raise


@app.post("/api/upload/confirm")
async def upload_confirm(session_id: str = Form(...), corrections: str = Form("{}")):
    import json
    session = _preview_sessions.get(session_id)
    if not session:
        return JSONResponse(status_code=404, content={"error": "Session expired or not found"})
    corrections_dict = json.loads(corrections) if corrections else {}
    result = json_confirm_upload(filepath=session["filepath"], date_str=session["date_str"], corrections=corrections_dict)
    _sync_json_to_sqlite(result)
    os.unlink(session["filepath"])
    del _preview_sessions[session_id]
    return {"status": "ok", "brands_count": len(result.get("brands", []))}


def _sync_json_to_sqlite(data: dict):
    """Sync JSON result into SQLite tables."""
    with db_session() as session:
        for brand_entry in data.get("brands", []):
            brand_name = brand_entry.get("brand", "")
            for item in brand_entry.get("items", []):
                item_name = item.get("name", "")
                pid = get_product_id(session, item_name)
                if pid is None:
                    prod = Product(name=item_name, brand=brand_name)
                    session.add(prod)
                    session.flush()
                    pid = prod.id
                for pp in item.get("prices", []):
                    upsert_price(session, pid, pp.get("price"), pp.get("date", ""), "excel_upload")


# ── Dashboard API (reads from SQLite) ────────────────────────

@app.get("/api/dashboard")
def get_dashboard():
    return {"brands": _db_to_brand_groups()}


# ── Backup / Restore API (SQLite .db files) ──────────────────

@app.get("/api/backups")
def list_backups():
    if not os.path.exists(BACKUP_DIR):
        return {"backups": []}
    files = sorted(glob.glob(os.path.join(BACKUP_DIR, "*.db")), reverse=True)
    backups = []
    for f in files:
        name = os.path.basename(f)
        backups.append({
            "name": name,
            "size": os.path.getsize(f),
            "time": datetime.fromtimestamp(os.path.getmtime(f), tz=BJT).strftime("%Y-%m-%d %H:%M:%S"),
        })
    return {"backups": backups}


@app.get("/api/backup/download/{backup_name:path}")
def download_backup(backup_name: str):
    backup_path = os.path.join(BACKUP_DIR, backup_name)
    if not os.path.exists(backup_path):
        return JSONResponse(status_code=404, content={"error": "Backup not found"})
    return FileResponse(backup_path, filename=backup_name, media_type="application/octet-stream")


@app.post("/api/backup")
def create_backup():
    from .database import DB_PATH, engine
    os.makedirs(BACKUP_DIR, exist_ok=True)
    ts = now_bj().strftime("%Y%m%d_%H%M%S")
    dst = os.path.join(BACKUP_DIR, f"backup_{ts}.db")
    if os.path.exists(DB_PATH):
        # Force WAL checkpoint so all data is in the main file
        with engine.connect() as conn:
            conn.execute(sa_text("PRAGMA wal_checkpoint(TRUNCATE)"))
        shutil.copy2(DB_PATH, dst)
        return {"status": "ok", "name": f"backup_{ts}.db"}
    return JSONResponse(status_code=404, content={"error": "No database to backup"})


@app.post("/api/restore/{backup_name:path}")
def restore_backup(backup_name: str):
    from .database import DB_PATH
    backup_path = os.path.join(BACKUP_DIR, backup_name)
    if not os.path.exists(backup_path):
        return JSONResponse(status_code=404, content={"error": "Backup not found"})
    # Close all connections by disposing engine
    from .database import engine
    engine.dispose()
    shutil.copy2(backup_path, DB_PATH)
    # Re-init
    from .database import init_db
    init_db()
    brands = _db_to_brand_groups()
    return {"status": "ok", "brands_count": len(brands)}


@app.post("/api/seed")
def trigger_seed():
    """Force re-seed products from template (idempotent)."""
    from .migration import seed_on_startup
    ok = seed_on_startup()
    from .database import db_session, Product
    with db_session() as s:
        count = s.query(Product).count()
        return {"status": "ok" if ok else "skipped", "products": count}


@app.post("/api/aliases/generate")
def generate_aliases():
    """Generate auto-aliases for products that don't have any yet."""
    from .seed_products import generate_aliases as gen
    from .database import db_session, ProductAlias
    with db_session() as s:
        count = gen(s)
    total = 0
    with db_session() as s:
        total = s.query(ProductAlias).count()
    return {"status": "ok", "generated": count, "total_aliases": total}


@app.post("/api/aliases/import")
def import_aliases():
    """Import aliases from user-provided JSON via file upload."""
    import tempfile, json
    file = UploadFile(...)  # placeholder — we'll read from disk instead
    return JSONResponse(status_code=400, content={"error": "Use POST /api/aliases/import/data with JSON body"})


@app.post("/api/aliases/import/data")
async def import_aliases_json(data: dict = {}):
    """Import aliases from JSON: [{\"standard\":\"云端之上\",\"alias\":\"云段之上门票\"}]"""
    from .seed_products import import_aliases as do_import
    items = data.get("aliases", [])
    if not items:
        return JSONResponse(status_code=400, content={"error": "No aliases in request"})
    with db_session() as s:
        count = do_import(s, items)
    total = 0
    with db_session() as s:
        from .database import ProductAlias
        total = s.query(ProductAlias).count()
    return {"status": "ok", "imported": count, "total_aliases": total}


@app.post("/api/aliases/import-from-template")
def import_aliases_from_template():
    """Import aliases from the template Excel file's column B."""
    import openpyxl
    from .seed_products import import_aliases as do_import

    KNOWN_BRANDS = {'云南','浙江','上海','湖北','湖南','河南','河北','广西',
        '江苏','内蒙','山东','江西','贵州','四川','福建','吉林',
        '广东','安徽','陕西','重庆','甘肃','哈尔滨','公司进口'}

    filepath = "/app/price_template.xlsx"
    if not os.path.exists(filepath):
        return JSONResponse(status_code=404, content={"error": "Template not found at /app/price_template.xlsx"})

    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    alias_list = []

    for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
        a, b = row[0], row[1]
        if a is None: continue
        text = str(a).strip()
        if not text or text in KNOWN_BRANDS: continue
        if b is not None and not isinstance(b, (int, float)):
            alias_text = str(b).strip()
            if alias_text and alias_text != text:
                alias_list.append({"standard": text, "alias": alias_text})

    with db_session() as s:
        count = do_import(s, alias_list)

    total = 0
    with db_session() as s:
        from .database import ProductAlias
        total = s.query(ProductAlias).count()
    return {"status": "ok", "found": len(alias_list), "imported": count, "total_aliases": total}


@app.post("/api/recover")
def recover_from_json():
    """Recover data from prices.json.bak into SQLite."""
    from .database import DB_PATH, init_db, engine
    import json
    bak_path = os.path.join(DATA_DIR, "prices.json.bak")
    if not os.path.exists(bak_path):
        return JSONResponse(status_code=404, content={"error": "No backup JSON found"})

    # Clear existing data
    engine.dispose()
    if os.path.exists(DB_PATH):
        os.remove(DB_PATH)
    init_db()

    with open(bak_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    with db_session() as session:
        for brand_entry in data.get("brands", []):
            brand_name = brand_entry.get("brand", "")
            for item in brand_entry.get("items", []):
                item_name = item.get("name", "")
                pid = get_product_id(session, item_name)
                if pid is None:
                    prod = Product(name=item_name, brand=brand_name)
                    session.add(prod)
                    session.flush()
                    pid = prod.id
                for pp in item.get("prices", []):
                    upsert_price(session, pid, pp.get("price"), pp.get("date", ""), "recovery")

    from .migration import seed_on_startup
    seed_on_startup()
    return {"status": "ok", "message": "Data recovered from JSON backup"}


# ── Item detail API ──────────────────────────────────────────

@app.get("/api/item/{item_name:path}")
def get_item(item_name: str):
    results = []
    with db_session() as session:
        prod = session.query(Product).filter(Product.name == item_name).first()
        if prod:
            records = session.query(PriceHistory).filter(
                PriceHistory.product_id == prod.id
            ).order_by(PriceHistory.price_date).all()
            prices = [{"date": r.price_date, "price": r.price} for r in records]
            results.append({"name": prod.name, "brand": prod.brand or "", "prices": prices})
    return {"items": results}


# ── Template download ────────────────────────────────────────

@app.get("/api/template")
def download_template():
    import openpyxl
    from openpyxl.styles import PatternFill, Font

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Template"
    date_str = now_bj().strftime("%Y%m%d")
    ws.cell(row=1, column=1, value=date_str)

    brands = _db_to_brand_groups()
    row_num = 2
    brand_fill = PatternFill(start_color="E6F0FF", end_color="E6F0FF", fill_type="solid")
    brand_font = Font(bold=True, color="1677FF")

    for brand_group in brands:
        ws.cell(row=row_num, column=1, value=brand_group.get("brand", ""))
        ws.cell(row=row_num, column=1).fill = brand_fill
        ws.cell(row=row_num, column=1).font = brand_font
        row_num += 1
        for item in brand_group.get("items", []):
            prices = item.get("prices", [])
            sorted_prices = sorted(
                [p for p in prices if p.get("price") is not None],
                key=lambda x: x.get("date", ""),
            )
            latest_price = sorted_prices[-1]["price"] if sorted_prices else None
            ws.cell(row=row_num, column=1, value=item.get("name", ""))
            if latest_price is not None:
                ws.cell(row=row_num, column=2, value=latest_price)
            row_num += 1

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 12
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=price_template_{date_str}.xlsx"},
    )


# ── 粘贴文本上传 + DeepSeek AI 比对 API ─────────────────

@app.post("/api/paste/preview")
async def paste_preview(data: dict = {}):
    """解析粘贴文本，返回匹配引擎结果。"""
    from .text_parser import parse_text
    from .matcher import load_match_data, match_item
    from .database import db_session
    text = data.get("text", "")
    if not text:
        return JSONResponse(status_code=400, content={"error": "No text provided"})
    items = parse_text(text)
    results = []
    with db_session() as s:
        md = load_match_data(s)
        for item in items:
            r = match_item(None, item["name"], md)
            # Get last price for deviation check
            last_price = None
            if r["product_id"]:
                from .database import PriceHistory
                last_record = s.query(PriceHistory).filter(
                    PriceHistory.product_id == r["product_id"]
                ).order_by(PriceHistory.price_date.desc()).first()
                if last_record:
                    last_price = last_record.price
            results.append({
                "input": item["name"],
                "matched_name": r["product_name"],
                "matched_id": r["product_id"],
                "brand": item.get("brand", ""),
                "price": item.get("price"),
                "last_price": last_price,
                "score": r["score"],
                "source": r["source"],
                "status": r["status"],
            })
    return {"items": results, "total": len(results)}


@app.post("/api/paste/deepseek-compare")
async def paste_deepseek_compare(data: dict = {}):
    """DeepSeek AI 比对。"""
    import requests, json, re
    items = data.get("items", [])
    if not items:
        return JSONResponse(status_code=400, content={"error": "No items"})

    from .database import db_session, Product, ProductAlias
    with db_session() as s:
        aliases = s.query(ProductAlias).order_by(ProductAlias.product_id, ProductAlias.sort_order).all()
        ref_lines = []
        seen = set()
        for a in aliases:
            if a.product_id not in seen:
                ref_lines.append(f"{a.alias} -> {a.product_id}")
                seen.add(a.product_id)

    prompt = f"""现有商品别名对照表（alias -> product_id）。匹配以下商品名称，返回JSON数组：
[{{"input":"原始名称","matched_name":"标准名称","matched_id":数字,"price":数字或null,"score":0-100}}]

规则：
- 优先精确匹配 alias
- 谐音/形近自动纠错
- score>=90自动通过，70-89需确认，<70不匹配

标准商品列表：
"""
    # Get product names
    with db_session() as s:
        products = s.query(Product).all()
        for p in products:
            prompt += f"  id={p.id}: {p.name}\n"

    prompt += "\n待匹配商品：\n"
    for item in items:
        prompt += f"\n{item.get('input','')}"

    try:
        resp = requests.post(
            "https://api.deepseek.com/chat/completions",
            headers={"Content-Type": "application/json",
                     "Authorization": "Bearer sk-270635d346f745b4871eab1c5f773e62"},
            json={"model": "deepseek-chat",
                  "messages": [{"role": "user", "content": prompt}],
                  "max_tokens": 4096, "temperature": 0.05},
            timeout=60
        )
        content = resp.json()["choices"][0]["message"]["content"]
        match = re.search(r'\[.*\]', content, re.DOTALL)
        ds_results = json.loads(match.group()) if match else []
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})

    return {"items": ds_results}


@app.post("/api/paste/confirm")
async def paste_confirm(data: dict = {}):
    """确认保存粘贴结果（含日期）。"""
    from .database import db_session, upsert_price
    items = data.get("items", [])
    price_date = data.get("price_date", "")
    if not price_date:
        price_date = now_bj().strftime("%Y-%m-%d")
    if not items:
        return JSONResponse(status_code=400, content={"error": "No items"})
    saved = 0
    with db_session() as s:
        for item in items:
            pid = item.get("matched_id")
            price = item.get("price")
            if pid and price is not None:
                upsert_price(s, pid, price, price_date, "text_paste")
                saved += 1
    return {"status": "ok", "saved": saved, "date": price_date}


# ── 别名管理 API ─────────────────────────────────────────

@app.get("/api/aliases/manage")
def list_aliases_manage():
    """列出所有标准商品及其别名（按 sort_order）。"""
    from .database import db_session, Product, ProductAlias
    result = []
    with db_session() as s:
        products = s.query(Product).order_by(Product.brand, Product.name).all()
        for prod in products:
            aliases = s.query(ProductAlias).filter(
                ProductAlias.product_id == prod.id
            ).order_by(ProductAlias.sort_order).all()
            alias_list = [{"id": a.id, "alias": a.alias, "source": a.source,
                          "sort_order": a.sort_order} for a in aliases]
            result.append({
                "product_id": prod.id,
                "name": prod.name,
                "brand": prod.brand or "",
                "aliases": alias_list,
            })
    return {"products": result}


@app.post("/api/aliases/reorder")
async def reorder_aliases(data: dict = {}):
    """保存别名排序。"""
    from .database import db_session, ProductAlias
    product_id = data.get("product_id")
    alias_ids = data.get("alias_ids", [])
    if not product_id or not alias_ids:
        return JSONResponse(status_code=400, content={"error": "Missing product_id or alias_ids"})
    with db_session() as s:
        for i, aid in enumerate(alias_ids):
            a = s.query(ProductAlias).filter(
                ProductAlias.id == aid,
                ProductAlias.product_id == product_id,
            ).first()
            if a:
                a.sort_order = i
    return {"status": "ok", "sorted": len(alias_ids)}


@app.post("/api/aliases/add")
async def add_alias(data: dict = {}):
    """手动添加别名。"""
    from .database import db_session, Product, ProductAlias
    product_id = data.get("product_id")
    alias_text = data.get("alias", "").strip()
    if not product_id or not alias_text:
        return JSONResponse(status_code=400, content={"error": "Missing fields"})
    with db_session() as s:
        existing = s.query(ProductAlias).filter(ProductAlias.alias == alias_text).first()
        if existing:
            return JSONResponse(status_code=400, content={"error": "Alias already exists"})
        max_order = s.query(func.max(ProductAlias.sort_order)).filter(
            ProductAlias.product_id == product_id
        ).scalar() or 0
        a = ProductAlias(product_id=product_id, alias=alias_text,
                        sort_order=max_order + 1, source="manual")
        s.add(a)
    return {"status": "ok", "alias_id": a.id}


@app.post("/api/aliases/delete")
async def delete_alias(data: dict = {}):
    """删除别名。"""
    from .database import db_session, ProductAlias
    alias_id = data.get("alias_id")
    if not alias_id:
        return JSONResponse(status_code=400, content={"error": "Missing alias_id"})
    with db_session() as s:
        a = s.query(ProductAlias).filter(ProductAlias.id == alias_id).first()
        if a:
            s.delete(a)
    return {"status": "ok", "deleted": alias_id}


@app.get("/api/products/search/{query:path}")
def search_products(query: str):
    """Search products by name (for autocomplete)."""
    from .database import db_session, Product
    if len(query) < 1:
        return {"results": []}
    results = []
    with db_session() as s:
        products = s.query(Product).filter(Product.name.contains(query)).limit(20).all()
        for p in products:
            results.append({"id": p.id, "name": p.name, "brand": p.brand or ""})
    return {"results": results}


@app.post("/api/aliases/learn")
async def learn_alias(data: dict = {}):
    """用户确认匹配后自动学习别名。"""
    from .database import db_session, Product, ProductAlias
    input_name = data.get("input", "").strip()
    product_id = data.get("product_id")
    if not input_name or not product_id:
        return JSONResponse(status_code=400, content={"error": "Missing input or product_id"})
    with db_session() as s:
        existing = s.query(ProductAlias).filter(ProductAlias.alias == input_name).first()
        if not existing:
            max_order = s.query(func.max(ProductAlias.sort_order)).filter(
                ProductAlias.product_id == product_id
            ).scalar() or 0
            a = ProductAlias(product_id=product_id, alias=input_name,
                            sort_order=max_order + 1, source="user_correction")
            s.add(a)
    return {"status": "ok", "learned": input_name}


@app.post("/api/aliases/edit-product")
async def edit_product_name(data: dict = {}):
    """修改标准商品名称。"""
    from .database import db_session, Product, ProductAlias
    product_id = data.get("product_id")
    new_name = data.get("name", "").strip()
    if not product_id or not new_name:
        return JSONResponse(status_code=400, content={"error": "Missing product_id or name"})
    with db_session() as s:
        prod = s.query(Product).filter(Product.id == product_id).first()
        if not prod:
            return JSONResponse(status_code=404, content={"error": "Product not found"})
        existing = s.query(Product).filter(Product.name == new_name, Product.id != product_id).first()
        if existing:
            return JSONResponse(status_code=400, content={"error": "Name already exists"})
        prod.name = new_name
    return {"status": "ok", "name": new_name}


@app.post("/api/aliases/edit-alias")
async def edit_alias(data: dict = {}):
    """编辑别名文本。"""
    from .database import db_session, ProductAlias
    alias_id = data.get("alias_id")
    new_alias = data.get("alias", "").strip()
    if not alias_id or not new_alias:
        return JSONResponse(status_code=400, content={"error": "Missing alias_id or alias"})
    with db_session() as s:
        a = s.query(ProductAlias).filter(ProductAlias.id == alias_id).first()
        if not a:
            return JSONResponse(status_code=404, content={"error": "Alias not found"})
        existing = s.query(ProductAlias).filter(ProductAlias.alias == new_alias, ProductAlias.id != alias_id).first()
        if existing:
            return JSONResponse(status_code=400, content={"error": "Alias already exists"})
        a.alias = new_alias
    return {"status": "ok", "alias": new_alias}


@app.post("/api/aliases/update-brand")
async def edit_product_brand(data: dict = {}):
    """修改商品品牌。"""
    from .database import db_session, Product
    product_id = data.get("product_id")
    new_brand = data.get("brand", "").strip()
    if not product_id or not new_brand:
        return JSONResponse(status_code=400, content={"error": "Missing fields"})
    with db_session() as s:
        prod = s.query(Product).filter(Product.id == product_id).first()
        if prod:
            prod.brand = new_brand
    return {"status": "ok"}


# ── Frontend static files ────────────────────────────────────

STATIC_DIR = Path(__file__).resolve().parent.parent / "static"
if STATIC_DIR.exists():
    assets_dir = STATIC_DIR / "assets"
    if assets_dir.exists():
        app.mount("/assets", StaticFiles(directory=str(assets_dir)), name="assets")
    @app.get("/")
    async def serve_root():
        index_path = STATIC_DIR / "index.html"
        if index_path.exists():
            return FileResponse(str(index_path))
        return JSONResponse(status_code=404, content={"detail": "Not Found"})
