"""
Excel Parser for price tracker — 价格追踪 Excel 解析 & Upsert 合并。

══════════════════════════════════════════════════════════════
两阶段上传流程：
  Phase 1 (parse_upload):  解析 Excel + 检测异常价格 → 返回预览，不保存
  Phase 2 (confirm_upload): 应用修正 → 真正合并保存到 prices.json
══════════════════════════════════════════════════════════════

日期格式说明：
  旧格式 "0621" → MMDD → 补2026 → 2026-06-21
  新格式 "260621" → YYMMDD → 2026-06-21
  统一归一化为 ISO (YYYY-MM-DD) 存储

Excel 布局：
  Row 1:    日期（col A）
  Row 2+:   品牌分区 → 商品行
    品牌行：col A = 品牌名（云南/浙江/上海…），col B 为空
    商品行：col A = 商品名，col B = 价格（int/float/空/0=无报价）

异常价格检测：对比最近相邻日期的价格，偏差 ≥ 20元 → 弹窗提醒
"""

from typing import Optional
import openpyxl
from .storage import load_all, save_all

# 已知品牌分组（区域级分组名，作为解析品牌的锚点）
KNOWN_BRANDS = [
    '云南', '浙江', '上海', '湖北', '湖南', '河南', '河北', '广西',
    '江苏', '内蒙', '山东', '江西', '贵州', '四川', '福建', '吉林',
    '广东', '安徽', '陕西', '重庆', '甘肃', '哈尔滨', '公司进口',
]

ALERT_THRESHOLD = 20  # 偏差 ≥ 20元 触发提醒


# ── 日期归一化 ──────────────────────────────────────────────

def normalize_date(raw: str) -> str:
    s = str(raw).strip().replace("-", "").replace("/", "")
    if len(s) == 4:          # MMDD → 补当前世纪 "26"
        return f"2026-{s[:2]}-{s[2:]}"
    if len(s) == 6:          # YYMMDD
        return f"20{s[:2]}-{s[2:4]}-{s[4:]}"
    if len(s) == 8:          # YYYYMMDD
        return f"{s[:4]}-{s[4:6]}-{s[6:]}"
    return str(raw).strip()


# ── 内部解析逻辑 ────────────────────────────────────────────

def _parse_excel_rows(filepath: str) -> tuple[str, list[dict], list[dict]]:
    """
    解析 Excel，返回 (date_str, new_brands, flat_item_entries)。
    不读写 prices.json，纯解析。
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=1, max_col=2, values_only=True))

    date_str = normalize_date(str(rows[0][0]).strip()) if rows and rows[0][0] else ""

    existing_data = load_all()
    existing_brands = existing_data.get("brands", [])

    brand_names = set(KNOWN_BRANDS)

    new_brands: list[dict] = []
    flat_entries: list[dict] = []
    current_brand: Optional[str] = None
    first_content_row = True

    def get_or_create_brand(name: str) -> dict:
        for b in new_brands:
            if b["brand"] == name:
                return b
        entry = {"brand": name, "items": []}
        new_brands.append(entry)
        return entry

    def find_item(brand_entry: dict, item_name: str) -> Optional[dict]:
        for it in brand_entry["items"]:
            if it["name"] == item_name:
                return it
        return None

    def get_historic_item(item_name: str) -> Optional[dict]:
        for b in existing_brands:
            for it in b.get("items", []):
                if it["name"] == item_name:
                    return {"name": it["name"], "prices": list(it["prices"])}
        return None

    for row in rows[1:]:
        col_a, col_b = row[0], row[1]
        if col_a is None:
            continue
        text = str(col_a).strip()
        if not text:
            continue

        is_brand = False
        if first_content_row:
            is_brand = True
            first_content_row = False
        elif col_b is None:
            is_brand = text in brand_names
        else:
            is_brand = False

        if is_brand:
            current_brand = text
            get_or_create_brand(current_brand)
        else:
            if current_brand is None:
                current_brand = text
                get_or_create_brand(current_brand)
                continue

            item_name = text
            price = col_b if (isinstance(col_b, (int, float)) and col_b > 0) else None

            brand_entry = get_or_create_brand(current_brand)
            item_entry = find_item(brand_entry, item_name)
            if item_entry is None:
                hist = get_historic_item(item_name)
                item_entry = hist if hist is not None else {"name": item_name, "prices": []}
                brand_entry["items"].append(item_entry)

            flat_entries.append({
                "brand": current_brand,
                "brand_entry": brand_entry,
                "item_entry": item_entry,
                "item_name": item_name,
                "price": price,
                "is_new": item_entry is None,
            })

    # Carry over historical items not in new upload
    for existing_brand in existing_brands:
        eb_name = existing_brand["brand"]
        cur = None
        for nb in new_brands:
            if nb["brand"] == eb_name:
                cur = nb
                break
        if cur is None:
            new_brands.append(dict(existing_brand))
        else:
            curr_names = {it["name"] for it in cur["items"]}
            for old_item in existing_brand.get("items", []):
                if old_item["name"] not in curr_names:
                    cur["items"].append(dict(old_item))

    return date_str, new_brands, flat_entries


def _get_last_price(item_entry: dict) -> Optional[float]:
    """Get the most recent non-None price from history (excluding current upload)."""
    valid = sorted(
        [p for p in item_entry.get("prices", []) if p.get("price") is not None],
        key=lambda x: x.get("date", ""),
    )
    if valid:
        return valid[-1]["price"]
    return None


# ── Phase 1: Preview ──────────────────────────────────────

def parse_upload(filepath: str) -> dict:
    """
    Phase 1: Parse Excel, detect price anomalies, return preview data.
    Does NOT write to prices.json.
    """
    date_str, new_brands, flat_entries = _parse_excel_rows(filepath)

    alerts = []

    for entry in flat_entries:
        price = entry["price"]
        if price is None:
            continue  # no price, no alert

        item_entry = entry["item_entry"]
        last_price = _get_last_price(item_entry)

        if last_price is None:
            continue  # no historical data, skip

        diff = abs(price - last_price)
        if diff >= ALERT_THRESHOLD:
            alerts.append({
                "item_name": entry["item_name"],
                "brand": entry["brand"],
                "old_price": last_price,
                "new_price": price,
                "diff": round(price - last_price, 1),
                "diff_abs": round(diff, 1),
            })

    total_items = sum(len(e["item_name"]) if isinstance(e, dict) else 0 for e in flat_entries)
    # More accurate: count unique item_names
    unique_items = len(set(e["item_name"] for e in flat_entries))

    return {
        "date_str": date_str,
        "parsed_data": {"brands": new_brands},
        "alerts": alerts,
        "total_items": unique_items,
    }


# ── Phase 2: Confirm & Save ───────────────────────────────

def confirm_upload(filepath: str, date_str: str, corrections: dict) -> dict:
    """
    Phase 2: Apply corrections and merge into prices.json.

    corrections: { "item_name": corrected_price_or_none }
      - corrected_price ∈ (0, ∞): use this price
      - corrected_price is None or 0: treat as "no quote"
    """
    # Re-parse to get the same structure (fast since we already have the data pattern)
    _, new_brands, _ = _parse_excel_rows(filepath)

    data = load_all()

    # Now do the actual upsert with corrections applied
    # We need to rebuild with corrections
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=1, max_col=2, values_only=True))

    existing_data = data
    existing_brands = existing_data.get("brands", [])
    brand_names = set(KNOWN_BRANDS)

    final_brands: list[dict] = []
    current_brand: Optional[str] = None
    first_content_row = True

    def goc_brand(name: str) -> dict:
        for b in final_brands:
            if b["brand"] == name:
                return b
        entry = {"brand": name, "items": []}
        final_brands.append(entry)
        return entry

    def find_item_fb(brand_entry: dict, item_name: str) -> Optional[dict]:
        for it in brand_entry["items"]:
            if it["name"] == item_name:
                return it
        return None

    def get_historic_item_fb(item_name: str) -> Optional[dict]:
        for b in existing_brands:
            for it in b.get("items", []):
                if it["name"] == item_name:
                    return {"name": it["name"], "prices": list(it["prices"])}
        return None

    for row in rows[1:]:
        col_a, col_b = row[0], row[1]
        if col_a is None:
            continue
        text = str(col_a).strip()
        if not text:
            continue

        is_brand = False
        if first_content_row:
            is_brand = True
            first_content_row = False
        elif col_b is None:
            is_brand = text in brand_names
        else:
            is_brand = False

        if is_brand:
            current_brand = text
            goc_brand(current_brand)
        else:
            if current_brand is None:
                current_brand = text
                goc_brand(current_brand)
                continue

            item_name = text
            raw_price = col_b if (isinstance(col_b, (int, float)) and col_b > 0) else None

            # Apply correction if exists
            if item_name in corrections:
                corr = corrections[item_name]
                if corr is not None and corr > 0:
                    price = corr
                else:
                    price = None
            else:
                price = raw_price

            brand_entry = goc_brand(current_brand)
            item_entry = find_item_fb(brand_entry, item_name)
            if item_entry is None:
                hist = get_historic_item_fb(item_name)
                item_entry = hist if hist is not None else {"name": item_name, "prices": []}
                brand_entry["items"].append(item_entry)

            _upsert_price(item_entry, date_str, price)

    # Carry over historical items
    for existing_brand in existing_brands:
        eb_name = existing_brand["brand"]
        cur = None
        for nb in final_brands:
            if nb["brand"] == eb_name:
                cur = nb
                break
        if cur is None:
            final_brands.append(dict(existing_brand))
        else:
            curr_names = {it["name"] for it in cur["items"]}
            for old_item in existing_brand.get("items", []):
                if old_item["name"] not in curr_names:
                    cur["items"].append(dict(old_item))

    data["brands"] = final_brands
    save_all(data)
    return data


def _upsert_price(item: dict, date_str: str, price: Optional[float]):
    for p in item["prices"]:
        if p["date"] == date_str:
            p["price"] = price
            return
    item["prices"].append({"date": date_str, "price": price})
